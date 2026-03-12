# ---
# jupyter:
#   kernelspec:
#     display_name: Python 3
#     language: python
#     name: python3
# ---

# %%
#|default_exp crosswalk




# %% [markdown]
# # Crosswalk: US O*NET SOC → UK SOC 2020
#
# This module builds a crosswalk from US O*NET SOC codes to UK SOC 2020 codes
# using ISCO-08 as a bridge:
#
# **O*NET SOC → US SOC 2018 → US SOC 2010 → ISCO-08 → UK SOC 2020**
# 

# %%
#|export
from pathlib import Path
import pandas as pd
import openpyxl

from aspectt_pipeline.const import DATA_DIR, ONET_DIR




# %% [markdown]
# ## Load UK SOC 2020 Framework
#
# The ONS SOC 2020 coding index Excel file contains a "SOC2020 framework" sheet
# with all 412 unit groups (4-digit codes) and their titles. This serves as the
# target classification system — every occupation in the final dataset corresponds
# to one of these unit groups.
# 

# %%
#|export
def load_uk_soc_framework(data_dir: Path = DATA_DIR) -> dict[int, str]:
    """Load UK SOC 2020 unit group codes and titles from the ONS coding index."""
    xlsx_path = data_dir / "soc2020volume2thecodingindexexcel03122025.xlsx"
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb['SOC2020 framework']
    groups = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[4] is not None:
            groups[int(row[4])] = str(row[5]).strip()
    wb.close()
    return groups




# %%
# Test
uk_groups = load_uk_soc_framework()
print(f"UK SOC 2020 unit groups: {len(uk_groups)}")
assert len(uk_groups) == 412




# %% [markdown]
# ## Load ISCO-08 → UK SOC 2020 Mapping
#
# The ONS coding index also contains ISCO-08 codes for each entry. We extract
# unique (ISCO-08, UK SOC 2020) pairs from the "SOC2020 coding index" sheet.
# This is the final leg of the crosswalk chain — it bridges the international
# standard to the UK-specific classification.
# 

# %%
#|export
def load_isco_to_uk_soc(data_dir: Path = DATA_DIR) -> pd.DataFrame:
    """Extract ISCO-08 → UK SOC 2020 mapping from the ONS coding index."""
    xlsx_path = data_dir / "soc2020volume2thecodingindexexcel03122025.xlsx"
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb['SOC2020 coding index']

    pairs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        isco, uk_soc = row[12], row[6]
        if isco is not None and uk_soc is not None:
            try:
                pairs.append({
                    'isco_08': str(int(isco)).zfill(4),
                    'uk_soc_2020': int(uk_soc),
                })
            except (ValueError, TypeError):
                continue
    wb.close()
    return pd.DataFrame(pairs).drop_duplicates()




# %%
isco_uk = load_isco_to_uk_soc()
print(f"ISCO → UK SOC unique pairs: {len(isco_uk)}")
print(f"Unique ISCO codes: {isco_uk['isco_08'].nunique()}")




# %% [markdown]
# ## Load BLS ISCO-08 ↔ US SOC 2010 Crosswalk
#
# The Bureau of Labor Statistics provides a crosswalk between ISCO-08 and
# US SOC 2010. We use the SOC 2010 → ISCO-08 direction to bridge from the
# US classification to the international standard.
# 

# %%
#|export
def load_isco_soc_crosswalk(data_dir: Path = DATA_DIR) -> pd.DataFrame:
    """Load BLS ISCO-08 ↔ US SOC 2010 crosswalk."""
    df = pd.read_excel(data_dir / "ISCO_SOC_Crosswalk.xls", header=5)
    df.columns = ['isco_08', 'isco_08_title', 'part', 'soc_2010', 'soc_2010_title', 'comment']
    df = df.dropna(subset=['isco_08', 'soc_2010'])
    df['isco_08'] = df['isco_08'].astype(str).str.strip().str.zfill(4)
    df['soc_2010'] = df['soc_2010'].astype(str).str.strip()
    return df[['isco_08', 'soc_2010']].drop_duplicates()




# %%
isco_soc = load_isco_soc_crosswalk()
print(f"ISCO ↔ SOC 2010: {len(isco_soc)} pairs")




# %% [markdown]
# ## Load BLS SOC 2010 ↔ SOC 2018 Crosswalk
#
# O*NET uses the SOC 2018 system, but the ISCO crosswalk is only available
# against SOC 2010. This BLS crosswalk bridges the two US SOC editions.
# It is a many-to-many mapping because occupations were split and merged
# between the 2010 and 2018 revisions.
# 

# %%
#|export
def load_soc_2010_to_2018(data_dir: Path = DATA_DIR) -> pd.DataFrame:
    """Load BLS SOC 2010 → SOC 2018 crosswalk."""
    df = pd.read_excel(data_dir / "soc_2010_to_2018_crosswalk.xlsx", header=7)
    df.columns = ['soc_2010', 'soc_2010_title', 'soc_2018', 'soc_2018_title']
    df = df.dropna(subset=['soc_2010', 'soc_2018'])
    df['soc_2010'] = df['soc_2010'].astype(str).str.strip()
    df['soc_2018'] = df['soc_2018'].astype(str).str.strip()
    return df[['soc_2010', 'soc_2018']].drop_duplicates()




# %%
soc_xwalk = load_soc_2010_to_2018()
print(f"SOC 2010 → 2018: {len(soc_xwalk)} pairs")




# %% [markdown]
# ## Load O*NET Occupations
#
# Load all 923 O*NET-SOC occupation codes and their titles. We also extract
# the 6-digit `base_soc` code (truncating the 2-digit suffix), which maps
# directly to the US SOC 2018 system — this is Step 1 of the crosswalk chain.
# 

# %%
#|export
def load_onet_occupations(onet_dir: Path = ONET_DIR) -> pd.DataFrame:
    """Load O*NET occupation codes and titles."""
    df = pd.read_csv(onet_dir / "Occupation Data.txt", sep='\t')
    df = df.rename(columns={'O*NET-SOC Code': 'onet_soc', 'Title': 'onet_title'})
    df['base_soc'] = df['onet_soc'].str[:7]
    return df[['onet_soc', 'onet_title', 'base_soc', 'Description']]




# %%
onet_occ = load_onet_occupations()
print(f"O*NET occupations: {len(onet_occ)}")




# %% [markdown]
# ## Build the Full Crosswalk
#
# Chains all four steps together: O*NET SOC → US SOC 2018 → US SOC 2010 →
# ISCO-08 → UK SOC 2020. The result is a table of (O*NET code, UK SOC code)
# pairs with uniform contribution weights: if an O*NET code maps to N
# distinct UK SOC codes, each receives weight 1/N.
# 

# %%
#|export
def build_crosswalk(data_dir: Path = DATA_DIR, onet_dir: Path = ONET_DIR) -> pd.DataFrame:
    """
    Build the full O*NET SOC → UK SOC 2020 crosswalk.

    Chain: O*NET SOC → SOC 2018 → SOC 2010 → ISCO-08 → UK SOC 2020

    Returns a DataFrame with columns:
    - onet_soc: O*NET-SOC code (e.g. '11-1011.00')
    - onet_title: O*NET occupation title
    - uk_soc_2020: UK SOC 2020 4-digit code
    - uk_soc_title: UK SOC 2020 unit group title
    - weight: contribution weight (1/N where N = number of UK SOC codes for this O*NET code)
    """
    # Load components
    onet_occ = load_onet_occupations(onet_dir)
    soc_2018_to_2010 = load_soc_2010_to_2018(data_dir).rename(
        columns={'soc_2010': 'soc_2010', 'soc_2018': 'soc_2018'}
    )
    # Reverse: SOC 2018 → SOC 2010
    soc_2018_to_2010_rev = soc_2018_to_2010[['soc_2018', 'soc_2010']].drop_duplicates()

    isco_soc = load_isco_soc_crosswalk(data_dir)
    # SOC 2010 → ISCO-08
    soc_to_isco = isco_soc[['soc_2010', 'isco_08']].drop_duplicates()

    isco_uk = load_isco_to_uk_soc(data_dir)
    uk_groups = load_uk_soc_framework(data_dir)

    # Chain joins
    # O*NET → SOC 2018 (base_soc matches soc_2018)
    xw = onet_occ[['onet_soc', 'onet_title', 'base_soc']].merge(
        soc_2018_to_2010_rev, left_on='base_soc', right_on='soc_2018', how='left'
    )
    # → SOC 2010 → ISCO-08
    xw = xw.merge(soc_to_isco, on='soc_2010', how='left')
    # → ISCO-08 → UK SOC 2020
    xw = xw.merge(isco_uk, on='isco_08', how='left')

    # Keep only matched rows
    xw = xw[xw['uk_soc_2020'].notna()].copy()
    xw['uk_soc_2020'] = xw['uk_soc_2020'].astype(int)
    xw['uk_soc_title'] = xw['uk_soc_2020'].map(uk_groups)

    # Deduplicate: unique (onet_soc, uk_soc_2020) pairs
    xw = xw[['onet_soc', 'onet_title', 'uk_soc_2020', 'uk_soc_title']].drop_duplicates()

    # Calculate weights: for each O*NET code, weight = 1/N (uniform across mapped UK codes)
    counts = xw.groupby('onet_soc')['uk_soc_2020'].transform('nunique')
    xw['weight'] = 1.0 / counts

    return xw.sort_values(['uk_soc_2020', 'onet_soc']).reset_index(drop=True)




# %%
crosswalk = build_crosswalk()
print(f"Crosswalk rows: {len(crosswalk)}")
print(f"O*NET codes covered: {crosswalk['onet_soc'].nunique()}")
print(f"UK SOC codes covered: {crosswalk['uk_soc_2020'].nunique()}")




# %%
# Example: show mappings for a specific UK SOC code
uk_code = 2134  # Programmers and software development professionals
subset = crosswalk[crosswalk['uk_soc_2020'] == uk_code]
print(f"\nUK SOC {uk_code}: {subset['uk_soc_title'].iloc[0] if len(subset) > 0 else 'N/A'}")
print(f"Mapped from {len(subset)} US O*NET codes:")
for _, row in subset.iterrows():
    print(f"  {row['onet_soc']} ({row['onet_title']}) - weight: {row['weight']:.3f}")




# %% [markdown]
# ## Reverse Mapping: UK SOC → O*NET Sources
#
# Inverts the crosswalk to get, for each UK SOC code, the list of contributing
# O*NET occupations and their weights. Used by the translation step to look up
# which US occupations feed into each UK occupation.
# 

# %%
#|export
def build_uk_to_onet_mapping(crosswalk: pd.DataFrame) -> dict[int, list[dict]]:
    """
    Build a reverse mapping: UK SOC → list of contributing O*NET codes with weights.

    Returns dict mapping uk_soc_2020 → [{onet_soc, onet_title, weight}, ...]
    """
    result = {}
    for uk_code, group in crosswalk.groupby('uk_soc_2020'):
        result[int(uk_code)] = [
            {
                'onet_soc': row['onet_soc'],
                'onet_title': row['onet_title'],
                'weight': row['weight'],
            }
            for _, row in group.iterrows()
        ]
    return result
